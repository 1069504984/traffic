# 学习如何进行对excel文件进行读取和写入
from openpyxl import load_workbook
from openpyxl import workbook
from openpyxl.styles import Font, PatternFill, colors
from openpyxl.styles.colors import RED, GREEN
from common.log_demo import logger

Fail_font = Font(color=RED)
Fail_pattern = PatternFill("solid", fgColor=colors.RED)
Pass_font = Font(color=GREEN)
Pass_pattern = PatternFill("solid", fgColor=colors.GREEN)
from common.read_conf import ReadConf
from common import project_path


class GetTel:
    def get_tel(self, filename):
        '''方法的作用是获取excel文件中的设置的指定测试数据'''
        wb = load_workbook(filename)
        sheet = wb["faker_test"]
        tel = sheet.cell(1, 2).value
        wb.close()
        return tel

    def back_write(self, filename, value):
        wb = load_workbook(filename)
        sheet = wb["faker_test"]
        sheet.cell(1, 2).value = value
        wb.save(filename)
        wb.close()


class DoExcel(GetTel):
    def __init__(self, filename, sheetname):
        self.filename = filename
        self.sheetname = sheetname

    def read_data(self):
        global sheet
        global final_data
        tel = self.get_tel(project_path.test_case_path)
        case_id = ReadConf(project_path.conf_path, "Request", "case_id").get_str()
        try:
            wb = load_workbook(self.filename)
            print("当前excel所有表单名字{}".format(wb.sheetnames))
            sheet = wb[self.sheetname]

        except Exception as e:
            logger.error("文件打开异常")
        print(sheet.max_row)
        list_1 = []
        for i in range(2, sheet.max_row + 1):
            dict_1 = {}
            if sheet.cell(i, 1).value == "":
                break
            dict_1["case_id"] = sheet.cell(i, 1).value
            dict_1["modular"] = sheet.cell(i, 2).value
            dict_1["method"] = sheet.cell(i, 3).value
            dict_1["url"] = sheet.cell(i, 4).value
            dict_1["header"] = sheet.cell(i, 5).value
            dict_1["title"] = sheet.cell(i, 6).value
            if sheet.cell(i, 6).value.find("faker_test_data") != -1:  # 如果在param中找到了tel这个字符串就执行下面的语句
                dict_1["param"] = sheet.cell(i, 7).value.replace("faker_test_data", str(tel))
                self.back_write(project_path.case_path, tel + 1)
            else:
                dict_1["param"] = sheet.cell(i, 7).value
            dict_1["sql"] = sheet.cell(i, 8).value
            dict_1["expected"] = sheet.cell(i, 9).value
            dict_1["check"] = sheet.cell(i, 12).value
            list_1.append(dict_1)
        final_data = []  # 空的列表 储存最终的测试用例数据
        if case_id == "all":  # 获取所有的用例，否则如果是列表就获取列表中的指定ID的用例数据
            final_data = list_1
        else:
            for i in eval(case_id):
                final_data.append(list_1[i - 1])
        return final_data

    def write_data(self, row, test_result, result):
        wb = load_workbook(self.filename)
        sheet = wb[self.sheetname]
        sheet.cell(row, 10).value = test_result
        logger.info("写入test_result成功 ：{}".format(test_result))
        sheet.cell(row, 11).value = result
        logger.info("写入result成功 ：{}".format(result))
        if result == 'pass':
            sheet.cell(row, 11).fill = Pass_pattern
        else:
            sheet.cell(row, 11).fill = Fail_pattern
        wb.save(self.filename)
        wb.close()

    def ceeat_excel(self, new_filename):
        wb = workbook.Workbook()
        wb.save(new_filename)

    def analysis_check(self, api_no, api_name, api_check, response):
        param = api_check.spilt("==")[1]
        value = response
        for key in param:
            temp = value.get(key)
            value = temp




if __name__ == '__main__':
    data = DoExcel(project_path.case_path, "camera").read_data()
    print(data[0])
    print(data[1])
    print(data[2])
    print(data[3])

    print(len(data))
    # test_data = DoExcel(project_path.exampels_path, "recharge").read_data()
    # print(test_data)
